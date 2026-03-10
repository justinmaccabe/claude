"""
jobs/export_excel.py
Generates a formatted Excel report from the latest weekly run.

Usage:
    python3 -m jobs.export_excel
    python3 -m jobs.export_excel --portfolio-value 100000 --output ~/Desktop/portfolio.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── import portfolio logic ────────────────────────────────────────────────────
sys.path.insert(0, ".")
from config import TICKERS, DATA, CASH_RATE_REAL, MARKET_VOL_TARGET
from data.store import Store
from signals.trend import TrendSignal
from signals.carry import CarrySignal
from signals.regime import RegimeSignal
from portfolio.allocator import Allocator, SignalOverlay
from jobs.weekly import EXPECTED_RETURNS, VOLS, CORR_MATRIX, TICKERS_LIST

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "navy":        "0A0E17",
    "dark":        "0D1525",
    "mid":         "162030",
    "blue":        "4A9EDD",
    "green":       "10B981",
    "amber":       "F5D76E",
    "red":         "E05C3A",
    "purple":      "A855F7",
    "white":       "DDEEFF",
    "muted":       "4A6FA5",
    "core_blue":   "4A9EDD",
    "factor_gold": "F5D76E",
}

SLEEVE_COLOURS = {"core": "4A9EDD", "factor": "F5D76E"}


def _font(bold=False, size=11, colour="DDEEFF", name="Arial"):
    return Font(bold=bold, size=size, color=colour, name=name)


def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)


def _border(style="thin", colour="162030"):
    s = Side(style=style, color=colour)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _pct(val):
    return f"{val:.1f}%" if val is not None else "—"


def _write(ws, row, col, value, bold=False, size=11, colour="DDEEFF",
           fill=None, align="left", border=True, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=size, colour=colour)
    cell.alignment = _align(h=align)
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _border()
    if num_format:
        cell.number_format = num_format
    return cell


# ── Build data ────────────────────────────────────────────────────────────────

def build_portfolio_data(portfolio_value: float):
    store  = Store(DATA["db_path"])
    prices = store.read_prices(tickers=TICKERS_LIST + ["VGRO"])

    trend  = TrendSignal(prices)
    carry  = CarrySignal(prices, store=store)
    regime = RegimeSignal(store=store, prices=prices)

    trend_df   = trend.signal()
    carry_df   = carry.signal()
    regime_det = regime.detect()

    trend_signals = pd.Series(dtype=float)
    if not trend_df.empty:
        trend_signals = trend_df["signal"].reindex(TICKERS_LIST).fillna(0)

    carry_signals = pd.Series(dtype=float)
    if not carry_df.empty and "HYG  (credit spread)" in carry_df.index:
        carry_signals = pd.Series({"HYG": carry_df.loc["HYG  (credit spread)", "signal"]})

    overlay = SignalOverlay(
        trend_signals=trend_signals,
        carry_signals=carry_signals,
        regime_tilts=regime.tilts(),
    )

    allocator  = Allocator(EXPECTED_RETURNS, VOLS, CORR_MATRIX)
    base_w     = allocator.base_weights()
    target_w   = allocator.target_weights(overlay=overlay, levered=True)
    stats      = allocator.portfolio_stats(target_w)
    base_stats = allocator.portfolio_stats(base_w)

    current_w = pd.Series({t: 0.0 for t in TICKERS_LIST})
    trades    = allocator.trade_list(current_w, portfolio_value, overlay=overlay)

    positions = []
    for t in TICKERS_LIST:
        tw = float(target_w.get(t, 0))
        positions.append({
            "ticker":        t,
            "label":         TICKERS[t]["label"],
            "sleeve":        TICKERS[t]["sleeve"],
            "exp_return":    EXPECTED_RETURNS[t],
            "vol":           VOLS[t],
            "ind_sharpe":    round((EXPECTED_RETURNS[t] - CASH_RATE_REAL) / VOLS[t], 2),
            "weight":        tw,
            "dollar_value":  tw * portfolio_value,
            "trend_signal":  int(trend_signals.get(t, 0)) if t in trend_signals.index else None,
        })

    return {
        "positions":   positions,
        "stats":       stats,
        "base_stats":  base_stats,
        "regime":      regime_det.value,
        "trend_df":    trend_df,
        "carry_df":    carry_df,
        "trades":      trades,
        "prices":      prices,
        "target_w":    target_w,
        "timestamp":   datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def write_summary(wb, data, portfolio_value):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["blue"]

    # Column widths
    for col, w in [(1,22),(2,14),(3,14),(4,14),(5,14),(6,14),(7,18),(8,14),(9,14)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "PORTFOLIO SIGNAL REPORT  ·  AQR 2026 CAPITAL MARKET ASSUMPTIONS"
    c.font  = _font(bold=True, size=14, colour=C["white"])
    c.fill  = _fill(C["navy"])
    c.alignment = _align(h="center")

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"Generated: {data['timestamp']}  ·  Target Vol: {MARKET_VOL_TARGET}%  ·  Cash Rate: {CASH_RATE_REAL}%"
    c.font  = _font(size=10, colour=C["muted"])
    c.fill  = _fill(C["dark"])
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # ── Portfolio stats ───────────────────────────────────────────────────────
    row = 4
    stats_headers = ["METRIC", "VALUE", "VS BASE"]
    for col, h in enumerate(stats_headers, 1):
        _write(ws, row, col, h, bold=True, size=9, colour=C["muted"],
               fill=C["dark"], align="center")
    ws.row_dimensions[row].height = 16

    s, b = data["stats"], data["base_stats"]
    stat_rows = [
        ("Expected Return (real)",  f"{s['expected_return']:.2f}%",  f"{s['expected_return']-b['expected_return']:+.2f}%"),
        ("Volatility",              f"{s['volatility']:.1f}%",        f"{s['volatility']-b['volatility']:+.1f}%"),
        ("Sharpe Ratio",            f"{s['sharpe_ratio']:.3f}",       f"{s['sharpe_ratio']-b['sharpe_ratio']:+.3f}"),
        ("Gross Exposure",          f"{s['gross_exposure']:.2f}×",    "—"),
        ("Regime",                  data["regime"],                    "—"),
    ]
    for i, (label, val, delta) in enumerate(stat_rows):
        r = row + 1 + i
        _write(ws, r, 1, label, colour=C["white"], fill=C["dark"])
        delta_colour = C["green"] if "+" in str(delta) else C["red"] if "-" in str(delta) else C["muted"]
        _write(ws, r, 2, val, bold=True, colour=C["amber"], fill=C["dark"], align="center")
        _write(ws, r, 3, delta, colour=delta_colour, fill=C["dark"], align="center")
        ws.row_dimensions[r].height = 18

    # ── Positions table ───────────────────────────────────────────────────────
    row = 11
    headers = ["TICKER","NAME","SLEEVE","EXP REAL RET","VOL","IND SHARPE",
               "TARGET WEIGHT","$ VALUE","TREND"]
    for col, h in enumerate(headers, 1):
        _write(ws, row, col, h, bold=True, size=9, colour=C["muted"],
               fill=C["mid"], align="center")
    ws.row_dimensions[row].height = 16

    sorted_pos = sorted(data["positions"], key=lambda x: x["weight"], reverse=True)
    for i, p in enumerate(sorted_pos):
        r = row + 1 + i
        row_fill = C["dark"] if i % 2 == 0 else C["navy"]
        sleeve_col = C["core_blue"] if p["sleeve"] == "core" else C["factor_gold"]

        _write(ws, r, 1, p["ticker"],      bold=True,  colour=sleeve_col, fill=row_fill)
        _write(ws, r, 2, p["label"],                   colour=C["white"], fill=row_fill)
        _write(ws, r, 3, p["sleeve"].upper(),           colour=sleeve_col, fill=row_fill, align="center")
        _write(ws, r, 4, p["exp_return"]/100,          colour=C["green"], fill=row_fill, align="center", num_format="0.0%")
        _write(ws, r, 5, p["vol"]/100,                 colour=C["blue"],  fill=row_fill, align="center", num_format="0.0%")
        _write(ws, r, 6, p["ind_sharpe"],              colour=C["amber"], fill=row_fill, align="center", num_format="0.00")
        _write(ws, r, 7, p["weight"],                  bold=True, colour=C["white"], fill=row_fill, align="center", num_format="0.0%")
        _write(ws, r, 8, p["dollar_value"],            colour=C["white"], fill=row_fill, align="right",  num_format='$#,##0')

        sig = p["trend_signal"]
        sig_text  = "▲ BULL" if sig == 1 else "▼ BEAR" if sig == -1 else "—"
        sig_colour = C["green"] if sig == 1 else C["red"] if sig == -1 else C["muted"]
        _write(ws, r, 9, sig_text, colour=sig_colour, fill=row_fill, align="center")
        ws.row_dimensions[r].height = 18

    # Sleeve subtotals
    for sleeve in ["core", "factor"]:
        pts = [p for p in sorted_pos if p["sleeve"] == sleeve]
        r = row + len(sorted_pos) + 1 + (0 if sleeve == "core" else 1)
        total_w = sum(p["weight"] for p in pts)
        total_v = sum(p["dollar_value"] for p in pts)
        sc = C["core_blue"] if sleeve == "core" else C["factor_gold"]
        _write(ws, r, 1, f"{sleeve.upper()} TOTAL", bold=True, colour=sc, fill=C["mid"])
        _write(ws, r, 7, total_w, bold=True, colour=sc, fill=C["mid"], align="center", num_format="0.0%")
        _write(ws, r, 8, total_v, bold=True, colour=sc, fill=C["mid"], align="right", num_format='$#,##0')
        for col in [2,3,4,5,6,9]:
            _write(ws, r, col, "", fill=C["mid"])


# ── Sheet 2: Signals ──────────────────────────────────────────────────────────

def write_signals(wb, data):
    ws = wb.create_sheet("Signals")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["amber"]

    for col, w in [(1,22),(2,14),(3,12),(4,12),(5,30)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "SIGNAL DASHBOARD"
    c.font  = _font(bold=True, size=13, colour=C["white"])
    c.fill  = _fill(C["navy"])
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 26

    # Trend signals
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="TREND SIGNALS  (12-1 Momentum)")
    c.font = _font(bold=True, size=10, colour=C["blue"])
    c.fill = _fill(C["dark"])

    row = 4
    for col, h in enumerate(["TICKER","12M RETURN","Z-SCORE","PERCENTILE","SIGNAL"], 1):
        _write(ws, row, col, h, bold=True, size=9, colour=C["muted"], fill=C["mid"], align="center")

    if not data["trend_df"].empty:
        for i, (ticker, trow) in enumerate(data["trend_df"].iterrows()):
            r = row + 1 + i
            rf = C["dark"] if i % 2 == 0 else C["navy"]
            sig = int(trow["signal"])
            sig_text   = "▲  BULLISH" if sig > 0 else "▼  BEARISH"
            sig_colour = C["green"] if sig > 0 else C["red"]
            _write(ws, r, 1, ticker,                    bold=True, colour=C["white"],  fill=rf)
            _write(ws, r, 2, trow["momentum_12_1"],               colour=C["amber"],  fill=rf, align="center", num_format="+0.0%;-0.0%;-")
            _write(ws, r, 3, trow["z_score"],                      colour=C["blue"],   fill=rf, align="center", num_format="+0.00;-0.00;-")
            _write(ws, r, 4, trow["percentile"]/100,               colour=C["muted"],  fill=rf, align="center", num_format="0%")
            _write(ws, r, 5, sig_text, bold=True,                  colour=sig_colour,  fill=rf, align="center")
            ws.row_dimensions[r].height = 18

    # Carry + Regime
    row = 4 + len(data["trend_df"]) + 3
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="CARRY SIGNALS")
    c.font = _font(bold=True, size=10, colour=C["purple"])
    c.fill = _fill(C["dark"])

    if not data["carry_df"].empty:
        for i, (name, crow) in enumerate(data["carry_df"].iterrows()):
            r = row + 1 + i
            rf = C["dark"] if i % 2 == 0 else C["navy"]
            sig = int(crow["signal"]) if pd.notna(crow["signal"]) else 0
            sig_text   = "▲  POSITIVE" if sig > 0 else "▼  NEGATIVE" if sig < 0 else "──  NEUTRAL"
            sig_colour = C["green"] if sig > 0 else C["red"] if sig < 0 else C["muted"]
            _write(ws, r, 1, name,       colour=C["white"],   fill=rf)
            _write(ws, r, 2, sig_text,   colour=sig_colour,   fill=rf, align="center", bold=True)
            _write(ws, r, 3, str(crow.get("note",""))[:40], colour=C["muted"], fill=rf)
            ws.row_dimensions[r].height = 18

    # Regime
    row = row + len(data["carry_df"]) + 3
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=f"REGIME:  {data['regime']}")
    regime_colour = C["green"] if data["regime"]=="RISK_ON" else C["red"] if data["regime"]=="RISK_OFF" else C["amber"]
    c.font = _font(bold=True, size=12, colour=regime_colour)
    c.fill = _fill(C["dark"])
    ws.row_dimensions[row].height = 24


# ── Sheet 3: Trade List ───────────────────────────────────────────────────────

def write_trades(wb, data, portfolio_value):
    ws = wb.create_sheet("Trade List")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["green"]

    for col, w in [(1,12),(2,22),(3,16),(4,16),(5,12),(6,16)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"TRADE LIST  ·  Portfolio Value: ${portfolio_value:,.0f}"
    c.font  = _font(bold=True, size=13, colour=C["white"])
    c.fill  = _fill(C["navy"])
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 26

    row = 3
    for col, h in enumerate(["TICKER","NAME","CURRENT WT","TARGET WT","ACTION","TRADE VALUE"], 1):
        _write(ws, row, col, h, bold=True, size=9, colour=C["muted"], fill=C["mid"], align="center")

    trades_sorted = data["trades"].sort_values("drift", ascending=False)
    for i, (ticker, trow) in enumerate(trades_sorted.iterrows()):
        r = row + 1 + i
        rf = C["dark"] if i % 2 == 0 else C["navy"]
        action = trow["action"]
        action_colour = C["green"] if action=="BUY" else C["red"] if action=="SELL" else C["muted"]
        label = TICKERS.get(ticker, {}).get("label", "")
        _write(ws, r, 1, ticker,                   bold=True, colour=C["white"],   fill=rf)
        _write(ws, r, 2, label,                               colour=C["muted"],   fill=rf)
        _write(ws, r, 3, trow["current_weight"],              colour=C["muted"],   fill=rf, align="center", num_format="0.0%")
        _write(ws, r, 4, trow["target_weight"],               colour=C["white"],   fill=rf, align="center", num_format="0.0%")
        _write(ws, r, 5, action,            bold=True,         colour=action_colour,fill=rf, align="center")
        _write(ws, r, 6, trow["trade_value"],                  colour=action_colour,fill=rf, align="right",  num_format='$#,##0')
        ws.row_dimensions[r].height = 18

    # Total row
    r = row + len(trades_sorted) + 1
    total_buys  = trades_sorted[trades_sorted["action"]=="BUY"]["trade_value"].sum()
    total_sells = trades_sorted[trades_sorted["action"]=="SELL"]["trade_value"].sum()
    _write(ws, r, 1, "TOTAL BUYS",  bold=True, colour=C["green"], fill=C["mid"])
    _write(ws, r, 6, total_buys,    bold=True, colour=C["green"], fill=C["mid"], align="right", num_format='$#,##0')
    _write(ws, r+1, 1, "TOTAL SELLS", bold=True, colour=C["red"], fill=C["mid"])
    _write(ws, r+1, 6, total_sells,   bold=True, colour=C["red"], fill=C["mid"], align="right", num_format='$#,##0')


# ── Sheet 4: Returns Data (for charting) ─────────────────────────────────────

def write_returns(wb, data):
    ws = wb.create_sheet("Returns Data")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["purple"]

    prices = data["prices"]
    target_w = data["target_w"]

    available = [t for t in TICKERS_LIST if t in prices.columns]
    w_norm = target_w.reindex(available).fillna(0)
    w_norm = w_norm / w_norm.sum()

    px_ret    = prices[available].pct_change().dropna()
    port_ret  = px_ret.dot(w_norm)
    port_cum  = (1 + port_ret).cumprod()

    headers = ["Date", "Portfolio"]
    has_bench = "VGRO" in prices.columns
    if has_bench:
        bench_ret = prices["VGRO"].pct_change().dropna()
        common    = port_cum.index.intersection(bench_ret.index)
        bench_cum = (1 + bench_ret.loc[common]).cumprod()
        headers.append("VGRO Benchmark")
    else:
        common = port_cum.index

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _font(bold=True, size=9, colour=C["muted"])
        c.fill = _fill(C["mid"])

    for i, date in enumerate(common):
        r = i + 2
        ws.cell(row=r, column=1, value=date.strftime("%Y-%m-%d"))
        ws.cell(row=r, column=2, value=round(float(port_cum.loc[date]), 4))
        if has_bench:
            ws.cell(row=r, column=3, value=round(float(bench_cum.loc[date]), 4))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    if has_bench:
        ws.column_dimensions["C"].width = 18

    # Add chart
    n_rows = len(common)
    chart = BarChart()
    from openpyxl.chart import LineChart
    chart = LineChart()
    chart.title = "Cumulative Returns vs Benchmark"
    chart.style = 10
    chart.y_axis.title = "Cumulative Return"
    chart.x_axis.title = "Date"
    chart.height = 14
    chart.width  = 24

    port_data  = Reference(ws, min_col=2, min_row=1, max_row=n_rows+1)
    chart.add_data(port_data, titles_from_data=True)

    if has_bench:
        bench_data = Reference(ws, min_col=3, min_row=1, max_row=n_rows+1)
        chart.add_data(bench_data, titles_from_data=True)

    ws.add_chart(chart, "E3")


# ── Main ─────────────────────────────────────────────────────────────────────

def generate(portfolio_value: float = 100_000, output_path: str = None):
    print("Building portfolio data...")
    data = build_portfolio_data(portfolio_value)

    wb = Workbook()

    print("Writing Summary sheet...")
    write_summary(wb, data, portfolio_value)

    print("Writing Signals sheet...")
    write_signals(wb, data)

    print("Writing Trade List sheet...")
    write_trades(wb, data, portfolio_value)

    print("Writing Returns Data sheet...")
    write_returns(wb, data)

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"portfolio_report_{ts}.xlsx"

    wb.save(output_path)
    print(f"\n✓ Report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--portfolio-value", type=float, default=100_000)
    parser.add_argument("--output", type=str, default=None)
    args = parser.parse_args()
    generate(portfolio_value=args.portfolio_value, output_path=args.output)
